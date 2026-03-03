from flask import render_template, redirect, url_for, flash, abort, request
from flask_login import login_required, current_user
from . import bp
from .forms import CourseForm, CourseSettingsForm, EnrollByEmailForm
from ..decorators import role_required, require_course_access
from ..extensions import db
from ..models import Course, Enrollment, User, Group, GroupMember, Assignment
from ..audit import log_action

from openpyxl import load_workbook


@bp.get("")
@login_required
def list_courses():
    if current_user.role == "TEACHER":
        courses = Course.query.filter_by(teacher_id=current_user.id).order_by(Course.created_at.desc()).all()
    else:
        courses = (
            db.session.query(Course)
            .join(Enrollment, Enrollment.course_id == Course.id)
            .filter(Enrollment.user_id == current_user.id)
            .order_by(Course.created_at.desc())
            .all()
        )
    return render_template("courses/list.html", courses=courses)


@bp.get("/create")
@login_required
@role_required("TEACHER")
def create_course():
    form = CourseForm()
    # default constraints
    if form.max_group_size.data is None:
        form.max_group_size.data = 5
    if form.max_groups.data is None:
        form.max_groups.data = 0
    return render_template("courses/create_edit.html", form=form, mode="create")


@bp.post("/create")
@login_required
@role_required("TEACHER")
def create_course_post():
    form = CourseForm()
    if not form.validate_on_submit():
        flash("Dữ liệu không hợp lệ", "danger")
        return render_template("courses/create_edit.html", form=form, mode="create")

    c = Course(code=form.code.data.strip(), name=form.name.data.strip(), teacher_id=current_user.id,
              max_group_size=form.max_group_size.data, max_groups=form.max_groups.data)
    db.session.add(c)
    db.session.commit()
    log_action(current_user.id, "course.create", "Course", c.id, {"code": c.code})
    flash("Tạo lớp/môn thành công", "success")
    return redirect(url_for("courses.detail_course", course_id=c.id))


@bp.get("/<int:course_id>")
@login_required
def detail_course(course_id: int):
    require_course_access(course_id)

    course = Course.query.get_or_404(course_id)
    groups = Group.query.filter_by(course_id=course_id).order_by(Group.created_at.desc()).all()
    assignments = Assignment.query.filter_by(course_id=course_id).order_by(Assignment.created_at.desc()).all()

    # Teacher settings form (for course owner)
    settings_form = None
    if current_user.role == "TEACHER" and course.teacher_id == current_user.id:
        settings_form = CourseSettingsForm()
        settings_form.max_group_size.data = course.max_group_size
        settings_form.max_groups.data = course.max_groups

    # Student: find my current group (if any) in this course
    my_group = None
    if current_user.role == "STUDENT":
        my_group = (
            db.session.query(Group)
            .join(GroupMember, GroupMember.group_id == Group.id)
            .filter(Group.course_id == course_id, GroupMember.user_id == current_user.id)
            .first()
        )

    enroll_form = EnrollByEmailForm()
    enrolled_students = (
        db.session.query(User)
        .join(Enrollment, Enrollment.user_id == User.id)
        .filter(Enrollment.course_id == course_id, User.role == "STUDENT")
        .order_by(User.name.asc())
        .all()
    )

    return render_template(
        "courses/detail.html",
        course=course,
        groups=groups,
        assignments=assignments,
        enroll_form=enroll_form,
        enrolled_students=enrolled_students,
        settings_form=settings_form,
        my_group=my_group,
        groups_count=len(groups),
    )



@bp.post("/<int:course_id>/settings")
@login_required
@role_required("TEACHER")
def update_course_settings(course_id: int):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        abort(403)

    form = CourseSettingsForm()
    if not form.validate_on_submit():
        flash("Dữ liệu cài đặt không hợp lệ", "danger")
        return redirect(url_for("courses.detail_course", course_id=course_id))

    course.max_group_size = form.max_group_size.data
    course.max_groups = form.max_groups.data
    db.session.commit()

    log_action(current_user.id, "course.update_settings", "Course", course_id, {
        "max_group_size": course.max_group_size,
        "max_groups": course.max_groups,
    })

    flash("Đã cập nhật quy định nhóm cho lớp", "success")
    return redirect(url_for("courses.detail_course", course_id=course_id))


@bp.post("/<int:course_id>/enroll")
@login_required
@role_required("TEACHER")
def enroll_student(course_id: int):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        abort(403)

    form = EnrollByEmailForm()
    if not form.validate_on_submit():
        flash("Email không hợp lệ", "danger")
        return redirect(url_for("courses.detail_course", course_id=course_id))

    email = form.email.data.strip().lower()
    u = User.query.filter_by(email=email).first()
    if not u or u.role != "STUDENT":
        flash("Không tìm thấy sinh viên với email này", "warning")
        return redirect(url_for("courses.detail_course", course_id=course_id))

    existed = Enrollment.query.filter_by(course_id=course_id, user_id=u.id).first()
    if existed:
        flash("Sinh viên đã được thêm vào lớp", "info")
        return redirect(url_for("courses.detail_course", course_id=course_id))

    db.session.add(Enrollment(course_id=course_id, user_id=u.id))
    db.session.commit()
    log_action(current_user.id, "course.enroll", "Course", course_id, {"student_email": email})
    flash("Đã thêm sinh viên vào lớp", "success")
    return redirect(url_for("courses.detail_course", course_id=course_id))


@bp.post("/<int:course_id>/import-excel")
@login_required
@role_required("TEACHER")
def import_students_excel(course_id: int):
    """Teacher upload Excel danh sách lớp.

    File .xlsx (sheet đầu): bắt buộc cột `email`, tùy chọn `name`, `password`.

    - Nếu user chưa tồn tại -> tạo STUDENT (password theo file, nếu trống dùng 123456)
    - Enroll vào course

    Lưu ý: đây là chức năng "thêm học sinh vào lớp" theo danh sách Excel.
    """
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        abort(403)

    f = request.files.get("excel")
    if not f or not f.filename.lower().endswith(".xlsx"):
        flash("Vui lòng chọn file .xlsx", "warning")
        return redirect(url_for("courses.detail_course", course_id=course_id))

    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.active

    header = [str(c).strip().lower() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    def idx_of(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    i_email = idx_of("email", "e-mail", "mail")
    i_name = idx_of("name", "ten", "tên")
    i_pw = idx_of("password", "pass", "matkhau", "mật khẩu", "mat_khau")

    if i_email is None:
        flash("File Excel phải có cột 'email'", "danger")
        return redirect(url_for("courses.detail_course", course_id=course_id))

    created_users = 0
    enrolled = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_email = row[i_email] if i_email < len(row) else None
        if not raw_email:
            continue

        email = str(raw_email).strip().lower()
        name = None
        if i_name is not None and i_name < len(row) and row[i_name]:
            name = str(row[i_name]).strip()
        if not name:
            name = email.split("@")[0]

        password = None
        if i_pw is not None and i_pw < len(row) and row[i_pw]:
            password = str(row[i_pw]).strip()
        if not password:
            password = "123456"

        u = User.query.filter_by(email=email).first()
        if u is None:
            u = User(name=name, email=email, role="STUDENT")
            u.set_password(password)
            db.session.add(u)
            db.session.flush()
            created_users += 1
        else:
            # nếu email đã tồn tại, chỉ cập nhật name nếu đang trống
            if not u.name:
                u.name = name

        if Enrollment.query.filter_by(course_id=course_id, user_id=u.id).first() is None:
            db.session.add(Enrollment(course_id=course_id, user_id=u.id))
            enrolled += 1

    db.session.commit()

    log_action(current_user.id, "course.import_excel", "Course", course_id, {
        "created_users": created_users,
        "enrolled": enrolled,
    })

    flash(f"Import xong: tạo {created_users} user, enroll {enrolled} vào lớp.", "success")
    return redirect(url_for("courses.detail_course", course_id=course_id))
