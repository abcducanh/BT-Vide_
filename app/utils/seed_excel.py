import os
from typing import List, Dict, Optional

from openpyxl import load_workbook

from ..extensions import db
from ..models import User


def _parse_users_xlsx(path: str, default_role: str) -> List[Dict]:
    """Read users from an .xlsx file (first sheet).

    Expected columns (case-insensitive):
      - email (required)
      - name (optional)
      - password (optional)
      - role (optional)

    Returns list of dicts: {email, name, password, role}
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header = [str(c).strip().lower() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    def idx_of(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    i_email = idx_of("email", "e-mail", "mail")
    i_name = idx_of("name", "ten", "tên")
    i_pw = idx_of("password", "pass", "matkhau", "mật khẩu", "mat_khau")
    i_role = idx_of("role")

    if i_email is None:
        raise ValueError(f"File {os.path.basename(path)} phải có cột 'email'.")

    items: List[Dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_email = row[i_email] if i_email < len(row) else None
        if not raw_email:
            continue
        email = str(raw_email).strip().lower()

        name = None
        if i_name is not None and i_name < len(row) and row[i_name]:
            name = str(row[i_name]).strip()
        if not name:
            name = email.split("@")[0]

        password = None
        if i_pw is not None and i_pw < len(row) and row[i_pw]:
            password = str(row[i_pw]).strip()

        role = default_role
        if i_role is not None and i_role < len(row) and row[i_role]:
            role = str(row[i_role]).strip().upper()

        items.append({"email": email, "name": name, "password": password, "role": role})

    return items


def seed_users_from_excel(
    teacher_xlsx_path: str,
    students_xlsx_path: Optional[str] = None,
    default_password: str = "123456",
) -> Dict[str, int]:
    """Idempotent seed users into DB from Excel files.

    - teacher_xlsx_path: one or many teacher rows (role default TEACHER)
    - students_xlsx_path: optional students list (role default STUDENT)

    If a user exists, we do not overwrite password. We only fill missing name.

    Returns counters: created, updated_name, skipped
    """
    created = 0
    updated_name = 0
    skipped = 0

    def upsert(items: List[Dict]):
        nonlocal created, updated_name, skipped
        for it in items:
            email = it["email"]
            name = it.get("name")
            password = it.get("password") or default_password
            role = it.get("role") or "STUDENT"

            u = User.query.filter_by(email=email).first()
            if u is None:
                u = User(name=name or email.split("@")[0], email=email, role=role)
                u.set_password(password)
                db.session.add(u)
                created += 1
            else:
                # do not overwrite password; only fill name if missing
                if not u.name and name:
                    u.name = name
                    updated_name += 1
                skipped += 1

    teachers = _parse_users_xlsx(teacher_xlsx_path, default_role="TEACHER")
    upsert(teachers)

    if students_xlsx_path:
        students = _parse_users_xlsx(students_xlsx_path, default_role="STUDENT")
        upsert(students)

    db.session.commit()

    return {"created": created, "updated_name": updated_name, "skipped": skipped}
